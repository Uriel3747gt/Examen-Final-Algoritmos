#Gencer Uriel Cortéz Ramírez 0907-22-10331
import os
from openpyxl import Workbook, load_workbook

archivo_vehiculos = 'vehiculos.xlsx'

def crear_vehiculo():
    marca = input("Ingrese la marca del vehículo: ")
    modelo = input("Ingrese el modelo del vehículo: ")
    año = input("Ingrese el año del vehículo: ")
    color = input("Ingrese el color del vehículo: ")

    nuevo_vehiculo = [marca, modelo, año, color]

    if not os.path.isfile(archivo_vehiculos):
        wb = Workbook()
        ws = wb.active
        ws.append(['Marca', 'Modelo', 'Año', 'Color'])
        ws.append(nuevo_vehiculo)
        wb.save(archivo_vehiculos)
    else:
        wb = load_workbook(archivo_vehiculos)
        ws = wb.active
        ws.append(nuevo_vehiculo)
        wb.save(archivo_vehiculos)

    print("Vehículo creado exitosamente.")

def editar_vehiculo():
    listar_vehiculos()
    indice = int(input("Ingrese el índice del vehículo que desea editar: ")) + 1

    wb = load_workbook(archivo_vehiculos)
    ws = wb.active

    if indice < 2 or indice > ws.max_row:
        print("Índice no válido.")
        return

    columna = input("Ingrese la columna que desea editar (A, B, C, D): ")
    nuevo_valor = input(f"Ingrese el nuevo valor para {columna}: ")

    ws[f"{columna}{indice}"] = nuevo_valor
    wb.save(archivo_vehiculos)

    print("Vehículo editado exitosamente.")

def eliminar_vehiculo():
    listar_vehiculos()
    indice = int(input("Ingrese el índice del vehículo que desea eliminar: ")) + 1

    wb = load_workbook(archivo_vehiculos)
    ws = wb.active

    if indice < 2 or indice > ws.max_row:
        print("Índice no válido.")
        return

    ws.delete_rows(indice)
    wb.save(archivo_vehiculos)

    print("Vehículo eliminado exitosamente.")

def listar_vehiculos():
    if not os.path.isfile(archivo_vehiculos):
        print("No hay vehículos registrados.")
    else:
        wb = load_workbook(archivo_vehiculos)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            print(row)

def carga_masiva():
    archivo_carga_masiva = input("Ingrese el nombre del archivo de carga masiva (formato texto separado por '|'): ")

    try:
        with open(archivo_carga_masiva, 'r') as file:
            lines = file.readlines()
            for line in lines:
                valores = line.strip().split('|')
                nuevo_vehiculo = [valores[0], valores[1], valores[2], valores[3]]

                if not os.path.isfile(archivo_vehiculos):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['Marca', 'Modelo', 'Año', 'Color'])
                    ws.append(nuevo_vehiculo)
                    wb.save(archivo_vehiculos)
                else:
                    wb = load_workbook(archivo_vehiculos)
                    ws = wb.active
                    ws.append(nuevo_vehiculo)
                    wb.save(archivo_vehiculos)

    except FileNotFoundError:
        print("Archivo no encontrado.")
        return

    print("Carga masiva completada.")

def main():
    while True:
        print("\n1. Crear Vehículo\n2. Editar Vehículo\n3. Eliminar Vehículo\n4. Listar Vehículos\n5. Carga Masiva\n6. Salir")
        opcion = input("Seleccione una opción: ")

        if opcion == '1':
            crear_vehiculo()
        elif opcion == '2':
            editar_vehiculo()
        elif opcion == '3':
            eliminar_vehiculo()
        elif opcion == '4':
            listar_vehiculos()
        elif opcion == '5':
            carga_masiva()
        elif opcion == '6':
            break
        else:
            print("Opción no válida. Inténtelo de nuevo.")

if __name__ == "__main__":
    main()
